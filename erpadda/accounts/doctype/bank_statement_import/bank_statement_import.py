# Copyright (c) 2019, VMRaid Technologies and contributors
# For license information, please see license.txt


import csv
import json
import re

import vmraid
import openpyxl
from vmraid import _
from vmraid.core.doctype.data_import.data_import import DataImport
from vmraid.core.doctype.data_import.importer import Importer, ImportFile
from vmraid.utils.background_jobs import enqueue
from vmraid.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

INVALID_VALUES = ("", None)


class BankStatementImport(DataImport):
	def __init__(self, *args, **kwargs):
		super(BankStatementImport, self).__init__(*args, **kwargs)

	def validate(self):
		doc_before_save = self.get_doc_before_save()
		if (
			not (self.import_file or self.google_sheets_url)
			or (doc_before_save and doc_before_save.import_file != self.import_file)
			or (doc_before_save and doc_before_save.google_sheets_url != self.google_sheets_url)
		):

			template_options_dict = {}
			column_to_field_map = {}
			bank = vmraid.get_doc("Bank", self.bank)
			for i in bank.bank_transaction_mapping:
				column_to_field_map[i.file_field] = i.bank_transaction_field
			template_options_dict["column_to_field_map"] = column_to_field_map
			self.template_options = json.dumps(template_options_dict)

			self.template_warnings = ""

		self.validate_import_file()
		self.validate_google_sheets_url()

	def start_import(self):

		preview = vmraid.get_doc("Bank Statement Import", self.name).get_preview_from_template(
			self.import_file, self.google_sheets_url
		)

		if "Bank Account" not in json.dumps(preview["columns"]):
			vmraid.throw(_("Please add the Bank Account column"))

		from vmraid.core.page.background_jobs.background_jobs import get_info
		from vmraid.utils.scheduler import is_scheduler_inactive

		if is_scheduler_inactive() and not vmraid.flags.in_test:
			vmraid.throw(_("Scheduler is inactive. Cannot import data."), title=_("Scheduler Inactive"))

		enqueued_jobs = [d.get("job_name") for d in get_info()]

		if self.name not in enqueued_jobs:
			enqueue(
				start_import,
				queue="default",
				timeout=6000,
				event="data_import",
				job_name=self.name,
				data_import=self.name,
				bank_account=self.bank_account,
				import_file_path=self.import_file,
				google_sheets_url=self.google_sheets_url,
				bank=self.bank,
				template_options=self.template_options,
				now=vmraid.conf.developer_mode or vmraid.flags.in_test,
			)
			return True

		return False


@vmraid.whitelist()
def get_preview_from_template(data_import, import_file=None, google_sheets_url=None):
	return vmraid.get_doc("Bank Statement Import", data_import).get_preview_from_template(
		import_file, google_sheets_url
	)


@vmraid.whitelist()
def form_start_import(data_import):
	return vmraid.get_doc("Bank Statement Import", data_import).start_import()


@vmraid.whitelist()
def download_errored_template(data_import_name):
	data_import = vmraid.get_doc("Bank Statement Import", data_import_name)
	data_import.export_errored_rows()


def parse_data_from_template(raw_data):
	data = []

	for i, row in enumerate(raw_data):
		if all(v in INVALID_VALUES for v in row):
			# empty row
			continue

		data.append(row)

	return data


def start_import(
	data_import, bank_account, import_file_path, google_sheets_url, bank, template_options
):
	"""This method runs in background job"""

	update_mapping_db(bank, template_options)

	data_import = vmraid.get_doc("Bank Statement Import", data_import)
	file = import_file_path if import_file_path else google_sheets_url

	import_file = ImportFile("Bank Transaction", file=file, import_type="Insert New Records")

	data = parse_data_from_template(import_file.raw_data)

	if import_file_path:
		add_bank_account(data, bank_account)
		write_files(import_file, data)

	try:
		i = Importer(data_import.reference_doctype, data_import=data_import)
		i.import_data()
	except Exception:
		vmraid.db.rollback()
		data_import.db_set("status", "Error")
		vmraid.log_error(title=data_import.name)
	finally:
		vmraid.flags.in_import = False

	vmraid.publish_realtime("data_import_refresh", {"data_import": data_import.name})


def update_mapping_db(bank, template_options):
	bank = vmraid.get_doc("Bank", bank)
	for d in bank.bank_transaction_mapping:
		d.delete()

	for d in json.loads(template_options)["column_to_field_map"].items():
		bank.append("bank_transaction_mapping", {"bank_transaction_field": d[1], "file_field": d[0]})

	bank.save()


def add_bank_account(data, bank_account):
	bank_account_loc = None
	if "Bank Account" not in data[0]:
		data[0].append("Bank Account")
	else:
		for loc, header in enumerate(data[0]):
			if header == "Bank Account":
				bank_account_loc = loc

	for row in data[1:]:
		if bank_account_loc:
			row[bank_account_loc] = bank_account
		else:
			row.append(bank_account)


def write_files(import_file, data):
	full_file_path = import_file.file_doc.get_full_path()
	parts = import_file.file_doc.get_extension()
	extension = parts[1]
	extension = extension.lstrip(".")

	if extension == "csv":
		with open(full_file_path, "w", newline="") as file:
			writer = csv.writer(file)
			writer.writerows(data)
	elif extension == "xlsx" or "xls":
		write_xlsx(data, "trans", file_path=full_file_path)


def write_xlsx(data, sheet_name, wb=None, column_widths=None, file_path=None):
	# from xlsx utils with changes
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	ws = wb.create_sheet(sheet_name, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)

			clean_row.append(value)

		ws.append(clean_row)

	wb.save(file_path)
	return True


@vmraid.whitelist()
def upload_bank_statement(**args):
	args = vmraid._dict(args)
	bsi = vmraid.new_doc("Bank Statement Import")

	if args.company:
		bsi.update(
			{
				"company": args.company,
			}
		)

	if args.bank_account:
		bsi.update({"bank_account": args.bank_account})

	return bsi
